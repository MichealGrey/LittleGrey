from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference

from src.core.types import EXCEL_TOOL_SCHEMA, ToolResult
from src.tools.base import BaseTool


class ExcelTool(BaseTool):
    name = "excel_tool"
    description = "创建和编辑 Excel 文件"
    input_schema = EXCEL_TOOL_SCHEMA
    is_risky = True

    def execute(self, params: dict[str, Any]) -> ToolResult:
        action = params["action"]
        file_path = params["file_path"]

        try:
            if action == "create":
                return self._create(params)
            elif action == "read":
                return self._read(params)
            elif action == "write":
                return self._write(params)
            elif action == "add_chart":
                return self._add_chart(params)
            else:
                return ToolResult(success=False, message=f"未知操作: {action}")
        except Exception as e:
            return ToolResult(success=False, message=f"Excel操作失败: {e}")

    def _create(self, params: dict[str, Any]) -> ToolResult:
        file_path = Path(params["file_path"])
        file_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        sheet_name = params.get("sheet_name", "Sheet1")
        ws.title = sheet_name

        headers = params.get("headers")
        rows = params.get("rows")
        data = params.get("data")

        if headers:
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)

        if rows:
            start_row = 2 if headers else 1
            for row_idx, row_data in enumerate(rows, start_row):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        if data:
            for row_key, row_data in data.items():
                row_num = int(row_key) if isinstance(row_key, str) and row_key.isdigit() else row_key
                if isinstance(row_data, list):
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_num, column=col_idx, value=value)
                elif isinstance(row_data, dict):
                    for col_key, value in row_data.items():
                        col_num = int(col_key) if isinstance(col_key, str) and col_key.isdigit() else col_key
                        ws.cell(row=row_num, column=col_num, value=value)

        wb.save(str(file_path))
        return ToolResult(
            success=True,
            file_path=str(file_path),
            message=f"Excel文件已创建: {file_path}",
        )

    def _read(self, params: dict[str, Any]) -> ToolResult:
        file_path = Path(params["file_path"])
        if not file_path.exists():
            return ToolResult(success=False, message=f"文件不存在: {file_path}")

        wb = load_workbook(str(file_path), read_only=True)
        sheet_name = params.get("sheet_name")
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        rows_data = []
        for row in ws.iter_rows(values_only=True):
            rows_data.append([cell for cell in row])

        wb.close()
        return ToolResult(
            success=True,
            data={"rows": rows_data, "sheet": ws.title},
            message=f"读取 {len(rows_data)} 行数据",
        )

    def _write(self, params: dict[str, Any]) -> ToolResult:
        file_path = Path(params["file_path"])
        if not file_path.exists():
            return ToolResult(success=False, message=f"文件不存在: {file_path}")

        wb = load_workbook(str(file_path))
        sheet_name = params.get("sheet_name")
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        data = params.get("data", {})
        for row_key, row_data in data.items():
            row_num = int(row_key) if isinstance(row_key, str) and row_key.isdigit() else row_key
            if isinstance(row_data, list):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_idx, value=value)
            elif isinstance(row_data, dict):
                for col_key, value in row_data.items():
                    col_num = int(col_key) if isinstance(col_key, str) and col_key.isdigit() else col_key
                    ws.cell(row=row_num, column=col_num, value=value)

        wb.save(str(file_path))
        return ToolResult(
            success=True,
            file_path=str(file_path),
            message=f"Excel文件已更新: {file_path}",
        )

    def _add_chart(self, params: dict[str, Any]) -> ToolResult:
        file_path = Path(params["file_path"])
        if not file_path.exists():
            return ToolResult(success=False, message=f"文件不存在: {file_path}")

        wb = load_workbook(str(file_path))
        ws = wb.active

        chart_type = params.get("chart_type", "bar")
        chart_title = params.get("chart_title", "")
        data_range = params.get("data_range", "A1:B10")

        if chart_type == "bar":
            chart = BarChart()
        elif chart_type == "line":
            chart = LineChart()
        elif chart_type == "pie":
            chart = PieChart()
        else:
            return ToolResult(success=False, message=f"不支持的图表类型: {chart_type}")

        chart.title = chart_title

        min_col, min_row, max_col, max_row = self._parse_range(data_range)
        data_ref = Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)
        cats_ref = Reference(ws, min_col=1, min_row=min_row + 1, max_row=max_row)

        if chart_type == "pie":
            chart.add_data(data_ref, titles_from_data=True)
        else:
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)

        ws.add_chart(chart, "E2")
        wb.save(str(file_path))

        return ToolResult(
            success=True,
            file_path=str(file_path),
            message=f"已添加{chart_type}图表到 {file_path}",
        )

    def _parse_range(self, range_str: str) -> tuple[int, int, int, int]:
        import re
        match = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", range_str.upper())
        if not match:
            return 1, 1, 2, 10

        def col_to_num(col: str) -> int:
            num = 0
            for c in col:
                num = num * 26 + (ord(c) - ord("A") + 1)
            return num

        return (
            col_to_num(match.group(1)),
            int(match.group(2)),
            col_to_num(match.group(3)),
            int(match.group(4)),
        )

    def _risk_check(self, params: dict[str, Any]) -> tuple[bool, str]:
        action = params.get("action", "")
        file_path = params.get("file_path", "")
        if action in ("create", "write", "add_chart") and Path(file_path).exists():
            return True, f"文件 {file_path} 已存在，继续操作将覆盖原有内容"
        return False, ""
